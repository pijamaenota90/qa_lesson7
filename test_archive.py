import os
import zipfile
from io import TextIOWrapper
from pypdf import PdfReader
from openpyxl import load_workbook



def test_archive_exists(create_archive):
    assert os.path.exists(create_archive)
    assert zipfile.is_zipfile(create_archive)


def test_archive_contains_files(create_archive):
    with zipfile.ZipFile(create_archive, 'r') as zip_file:
        file_list = zip_file.namelist()

        assert 'test1.pdf' in file_list
        assert 'test2.xlsx' in file_list
        assert 'test3.csv' in file_list


def test_pdf_check(create_archive):
    with zipfile.ZipFile(create_archive, 'r') as zip_file:
        with zip_file.open('test1.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            first_page = reader.pages[0]
            text = first_page.extract_text()

            assert "SIPp reference documentation" in text


def test_xlsx_check(create_archive):
    with zipfile.ZipFile(create_archive, 'r') as zip_file:
        with zip_file.open('test2.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            assert sheet.cell(row=2, column=3).value == "test@mail.ru"


def test_csv_check(create_archive):
    with zipfile.ZipFile(create_archive, 'r') as zip_file:
        with zip_file.open('test3.csv') as csv_file:
            data = csv_file.read().decode('windows-1251')

            assert "123" in data
